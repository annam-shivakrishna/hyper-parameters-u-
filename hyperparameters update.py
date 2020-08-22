# -*- coding: utf-8 -*-
"""
Created on Sat Aug 22 12:32:34 2020

@author: ANNAM SHIVA KRISHNA
"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter

filename = "F:\\Analytics path notes\\untitled0.py"

def exists(var):
     return var in globals()
case = exists("hyperparameters") 

if case==True:
    print("ok")

"""
# Create new workbook
wb = Workbook()
sheet1 = wb.create_sheet('sheet1',0)

active = wb["sheet1"]

ws = wb.active
import datetime
#ws["A2"] = datetime.datetime.now()

# parameters

hyperparameters = [
    ['objective', 'binary'],
    ['metric', 'auc'],
    ['learning_rate', 1],
    ['max_depth', 1],
    ['num_leaves', 20],
    ['feature_fraction', 0.8],
    ['subsample', 0.2]
]          

for rows in hyperparameters:
    ws.append(rows)


"""

wb = xlsxwriter.Workbook('new excel2')
worksheet = wb.add_worksheet()

hyperparameters = [
    ['objective', 'binary'],
    ['metric', 'auc'],
    ['learning_rate', 1],
    ['max_depth', 1],
    ['num_leaves', 400],
    ['feature_fraction', 10],
    ['subsample', 0.3]
]        

row=0
col=0
for parameters,values in (hyperparameters):
    worksheet.write(row,col,parameters)
    worksheet.write(row,col+1,values)
    worksheet.write(row,col+2,values)
    row+=1


wb.close()
   

















